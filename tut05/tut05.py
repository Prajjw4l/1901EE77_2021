#Submitted by - Prajjwal Kumar 
#Roll no - 1901EE77

from openpyxl import Workbook
import csv
import os

student = {} #variable (dict) 

map = {"AA": 10,"AB": 9,"BB": 8,"BC": 7,"CC": 6,"CD": 5,"DD": 4,"DD*": 4,"F": 0,"F*": 0,"I": 0,} #mapping the grade to INT values

courses = {}

def calculate_SPI(grades, credit):  #utility fun to calculate CPI 
    avg = 0.0
    total_cred = 0
    for i in range(len(grades)):
        avg += grades[i]*credit[i]
    for c in credit:
        total_cred += c
    return round(avg/total_cred, 2)

def generate_marksheet():
    if os.path.exists("output") == False:
        os.makedirs("output")   #making new directry for output 
    with open("names-roll.csv", "r") as file: #opening the name - roll file 
        read_file = csv.DictReader(file)
        for i in read_file:
            student[i["Roll"]] = {"Name": i["Name"]}
    with open("subjects_master.csv", "r") as file: #opening the subjects_master file 
        read_file = csv.DictReader(file)
        for i in read_file:
            courses[i["subno"]] = {"subname": i["subname"], "ltp": i["ltp"], "crd": i["crd"]}
    
    with open("grades.csv", "r") as file:
        read_file = csv.DictReader(file)
        for i in read_file:
            try:
                student[i["Roll"]][i["Sem"]]
            except KeyError:
                student[i["Roll"]][i["Sem"]] = {}
            student[i["Roll"]][i["Sem"]][i["SubCode"]] = {"Grade": i["Grade"].strip(), "Sub_Type": i["Sub_Type"]}
    for i in student:
        work_book = Workbook()  #initialising the lists to store the data of each students
        sem_wise_credit = [0,0,0,0,0,0,0,0]
        total_credit = [0,0,0,0,0,0,0,0]
        SPI = [0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0]
        CPI = [0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0] 
        for j in range(1, 9):
            credit = []
            grades = []
            sheet = work_book.create_sheet("Sem" + str(j))
            sheet.append(["S.NO.","Sub No.","Name of Sub.","L-T-P","Credit of Sub.","Type of Subject","Grade of Sub.",])
            l = 1
            try:
                student[i][str(j)]
            except KeyError:
                continue
            for k in student[i][str(j)]:
                sheet.append([l,k,courses[k]["subname"],courses[k]["ltp"],courses[k]["crd"],student[i][str(j)][k]["Sub_Type"],student[i][str(j)][k]["Grade"]])
                credit.append(int(courses[k]["crd"]))
                grades.append(map[student[i][str(j)][k]["Grade"]])
                l += 1
            for c in credit:
                sem_wise_credit[j-1]+=c
            SPI[j-1]=calculate_SPI(grades, credit) #calling the utility function to calculate SPI or CPI 
            if j>1:
                total_credit[j-1]=total_credit[j-2]+sem_wise_credit[j-1]
                CPI[j-1]=calculate_SPI(SPI[:j],sem_wise_credit[:j])
            else:
                total_credit[j-1]=sem_wise_credit[j-1]
                CPI[j-1]=SPI[j-1]
        sheet = work_book.create_sheet("Overall")  #creating the required (overall)sheet with all details  
        sheet.append(["Roll", i])
        sheet.append(["Name of Student", student[i]["Name"]])
        sheet.append(["Branch", i[4:6]])
        sheet.append(["Sem. No.", 1,2,3,4,5,6,7,8])
        t = ["Semester wise Credit Taken"]  #temprory variable 
        for j in sem_wise_credit:
            t.append(str(j))
        sheet.append(t)
        t = ["SPI"]
        for j in SPI:
            t.append(str(j))
        sheet.append(t)
        t = ["Total Credits Taken"]
        for j in total_credit:
            t.append(str(j))
        sheet.append(t)
        t = ["CPI"]
        for j in CPI:
            t.append(str(j))
        sheet.append(t)
        del work_book["Sheet"] #deleting the blank sheet which was created by default 
        work_book.save(os.path.join("output", i + ".xlsx")) #saving the Output folder 

generate_marksheet()